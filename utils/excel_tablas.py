"""
Exporta las vistas de tabla a .xlsx usando openpyxl (gratuito).
  - Deudores por mes
  - Alumnos al día / en deuda
  - Pagos mensuales
"""
import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from utils.fecha import fmt_fecha
from utils.ui_helpers import dialogo_guardar_o_abrir, fmt_monto


# ── Paleta ───────────────────────────────────────────────────────────────
_COLOR_HEADER   = "2C3E50"
_COLOR_TH_FG    = "FFFFFF"
_COLOR_TH_BG    = "DDE3EA"
_COLOR_FILA_PAR = "F7F9FB"
_COLOR_TOTAL_BG = "D5F5E3"
_COLOR_VERDE    = "1E8449"
_COLOR_ROJO     = "C0392B"
_COLOR_NARANJA  = "E67E22"


# ── Estilos reutilizables ────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)


def _alinear(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ── Helpers internos ─────────────────────────────────────────────────────
def _nuevo_wb(titulo, subtitulo):
    """Crea un Workbook con una hoja activa y escribe el encabezado."""
    wb    = Workbook()
    ws    = wb.active
    ws.title = titulo[:31]

    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=10)
    celda           = ws.cell(1, 1)
    celda.font      = _font(bold=True, color=_COLOR_HEADER, size=13)
    celda.alignment = _alinear("center")

    ws.append([subtitulo])
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=10)
    celda2           = ws.cell(2, 1)
    celda2.font      = _font(color="555555", size=9)
    celda2.alignment = _alinear("center")

    ws.append([])

    return wb, ws


def _escribir_header(ws, fila, columnas):
    """Escribe la fila de encabezados con fondo oscuro."""
    for col_idx, texto in enumerate(columnas, start=1):
        c           = ws.cell(row=fila, column=col_idx, value=texto)
        c.font      = _font(bold=True, color=_COLOR_HEADER, size=9)
        c.fill      = _fill(_COLOR_TH_BG)
        c.alignment = _alinear("center")
        c.border    = _borde_fino()


def _escribir_fila(ws, fila, valores, es_par, formatos_der=None, color_override=None):
    """
    Escribe una fila de datos con colores alternados.
    formatos_der:  set de índices de columna (1-based) a alinear a la derecha.
    color_override: dict {col_idx: hex_color} para colorear texto de celdas específicas.
    """
    bg             = _COLOR_FILA_PAR if es_par else "FFFFFF"
    formatos_der   = formatos_der or set()
    color_override = color_override or {}

    for col_idx, valor in enumerate(valores, start=1):
        alin        = "right" if col_idx in formatos_der else "left"
        color       = color_override.get(col_idx, "000000")
        c           = ws.cell(row=fila, column=col_idx, value=valor)
        c.font      = _font(color=color, size=9)
        c.fill      = _fill(bg)
        c.alignment = _alinear(alin)
        c.border    = _borde_fino()


def _escribir_total(ws, fila, n_cols, col_label, col_valor, texto_label, texto_valor):
    """Escribe la fila de total con fondo verde claro."""
    for col_idx in range(1, n_cols + 1):
        valor = ""
        if col_idx == col_label:
            valor = texto_label
        elif col_idx == col_valor:
            valor = texto_valor
        c           = ws.cell(row=fila, column=col_idx, value=valor)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(_COLOR_TOTAL_BG)
        c.alignment = _alinear("right" if col_idx == col_valor else "left")
        c.border    = _borde_fino()


def _ajustar_anchos(ws, anchos):
    """anchos: lista de anchos en caracteres, uno por columna."""
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _ruta_tmp(nombre):
    return os.path.join(tempfile.gettempdir(), nombre)


# ════════════════════════════════════════════════════════════════════════
# 1. DEUDORES POR MES
# ════════════════════════════════════════════════════════════════════════
def exportar_excel_deudores(rows, mes_sel):
    """
    rows: lista de dicts con alumno_nombre, tutor_nombre,
          categoria_nombre, ultimo_pago
    mes_sel: string del mes (ej: "mayo")
    """
    fecha  = datetime.today().strftime("%d/%m/%Y %H:%M")
    titulo = f"Deudores — {mes_sel.capitalize()}"
    subtit = f"Generado el {fecha}  |  Total deudores: {len(rows)}"

    wb, ws = _nuevo_wb(titulo, subtit)

    headers = ["Alumno", "Tutor", "Categoría", "Último Pago de Cuota"]
    fila_h  = 4
    _escribir_header(ws, fila_h, headers)

    for i, r in enumerate(rows):
        fila_idx = fila_h + 1 + i
        valores  = [
            r.get("alumno_nombre")    or "—",
            r.get("tutor_nombre")     or "—",
            r.get("categoria_nombre") or "—",
            fmt_fecha(r.get("ultimo_pago")) or "Sin pagos registrados",
        ]
        _escribir_fila(ws, fila_idx, valores, es_par=(i % 2 == 0))

    _ajustar_anchos(ws, [38, 28, 22, 24])

    nombre = f"deudores_{mes_sel}.xlsx"
    ruta   = _ruta_tmp(nombre)
    wb.save(ruta)
    dialogo_guardar_o_abrir(ruta, nombre)


# ════════════════════════════════════════════════════════════════════════
# 2. ALUMNOS AL DÍA / EN DEUDA
# ════════════════════════════════════════════════════════════════════════
def exportar_excel_alumnos_al_dia(rows, mes_sel, estado_sel):
    """
    rows: lista de dicts enriquecidos con estado, alerta, ultimo_pago
    mes_sel: string del mes
    estado_sel: "Todos" | "Al día" | "En deuda"
    """
    fecha    = datetime.today().strftime("%d/%m/%Y %H:%M")
    al_dia   = sum(1 for r in rows if r.get("estado") == "Al día")
    en_deuda = sum(1 for r in rows if r.get("estado") == "En deuda")
    titulo   = f"Estado de Alumnos — {mes_sel.capitalize()}"
    subtit   = (f"Generado el {fecha}  |  "
                f"Al día: {al_dia}   En deuda: {en_deuda}   Total: {len(rows)}")

    wb, ws = _nuevo_wb(titulo, subtit)

    headers = ["Alumno", "Tutor", "Categoría", "Estado", "Último Pago", "Alerta"]
    fila_h  = 4
    _escribir_header(ws, fila_h, headers)

    for i, r in enumerate(rows):
        fila_idx   = fila_h + 1 + i
        estado     = r.get("estado") or "—"
        alerta_txt = "Sin pago reciente" if r.get("alerta") else "OK"
        valores    = [
            r.get("alumno_nombre")    or "—",
            r.get("tutor_nombre")     or "—",
            r.get("categoria_nombre") or "—",
            estado,
            fmt_fecha(r.get("ultimo_pago")) or "Sin pagos",
            alerta_txt,
        ]
        color_estado = _COLOR_VERDE if estado == "Al día" else _COLOR_ROJO
        color_alerta = _COLOR_NARANJA if r.get("alerta") else _COLOR_VERDE
        _escribir_fila(
            ws, fila_idx, valores, es_par=(i % 2 == 0),
            color_override={4: color_estado, 6: color_alerta}
        )

    _ajustar_anchos(ws, [36, 26, 20, 14, 18, 20])

    nombre = f"alumnos_estado_{mes_sel}.xlsx"
    ruta   = _ruta_tmp(nombre)
    wb.save(ruta)
    dialogo_guardar_o_abrir(ruta, nombre)


# ════════════════════════════════════════════════════════════════════════
# 3. PAGOS MENSUALES
# ════════════════════════════════════════════════════════════════════════
def exportar_excel_pagos_mensuales(rows, mes_val, anio_val, criterio_txt):
    """
    rows: lista de dicts de obtener_pagos_mensuales
    mes_val, anio_val: strings de los filtros aplicados
    criterio_txt: "Fecha de Emisión (Caja Real)" | "Mes de la Cuota (Devengado)"
    """
    fecha       = datetime.today().strftime("%d/%m/%Y %H:%M")
    grand_total = sum(float(r.get("monto") or 0) for r in rows)
    titulo      = "Pagos Mensuales"
    subtit      = (f"Generado el {fecha}  |  "
                   f"Mes: {mes_val}   Año: {anio_val}   Criterio: {criterio_txt}  |  "
                   f"Registros: {len(rows)}")

    wb, ws = _nuevo_wb(titulo, subtit)

    headers = ["Tutor", "Alumno", "Categoría", "Mes Cuota",
               "Fecha Emisión", "Monto ($)", "Forma de Pago", "Descripción"]
    n_cols  = len(headers)
    fila_h  = 4
    _escribir_header(ws, fila_h, headers)

    for i, r in enumerate(rows):
        fila_idx = fila_h + 1 + i
        monto    = float(r.get("monto") or 0)
        valores  = [
            r.get("tutor_nombre")     or "—",
            r.get("alumno_nombre")    or "—",
            r.get("categoria_nombre") or "—",
            (r.get("mes_pago") or "—").capitalize(),
            fmt_fecha(r.get("fecha_emision")) or "—",
            monto,
            (r.get("forma_pago") or "—").capitalize(),
            r.get("descripcion") or "—",
        ]
        _escribir_fila(ws, fila_idx, valores,
                       es_par=(i % 2 == 0), formatos_der={6})
        ws.cell(row=fila_idx, column=6).number_format = '$#,##0.00'

    fila_total = fila_h + 1 + len(rows)
    _escribir_total(ws, fila_total, n_cols,
                    col_label=5, col_valor=6,
                    texto_label="TOTAL",
                    texto_valor=grand_total)
    ws.cell(row=fila_total, column=6).number_format = '$#,##0.00'

    primera_fila_dato = fila_h + 1
    ultima_fila_dato  = fila_h + len(rows)
    if len(rows) > 0:
        col_monto_letra = get_column_letter(6)
        ws.cell(row=fila_total, column=6).value = (
            f"=SUM({col_monto_letra}{primera_fila_dato}"
            f":{col_monto_letra}{ultima_fila_dato})"
        )

    _ajustar_anchos(ws, [26, 32, 18, 13, 16, 13, 16, 28])

    nombre = "pagos_mensuales.xlsx"
    ruta   = _ruta_tmp(nombre)
    wb.save(ruta)
    dialogo_guardar_o_abrir(ruta, nombre)